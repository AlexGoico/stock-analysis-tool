from dotenv import load_dotenv
load_dotenv()

import os
from time import sleep
import argparse
from datetime import datetime

from AlphaVantage_Backend import AlphaVantage_Backend
from Metrics import Metrics

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_CURRENCY_USD
from openpyxl.utils.dataframe import dataframe_to_rows

av_key = os.getenv("ALPHA_VANTAGE")

backend = AlphaVantage_Backend(av_key)
metrics = Metrics(backend)


def analyze_ticker(ticker):
  assets = metrics.current_assets(ticker)
  liabilities = metrics.current_liabilities(ticker)
  al_ratio = {dt: total_assets / liabilities[dt] for dt, total_assets in assets.items()}

  metrics5years = pd.DataFrame([
    metrics.revenues(ticker),
    metrics.profits(ticker),
    metrics.net_profit_margins(ticker),
    al_ratio
  ], index=['revenues', 'net_profits', 'net_profit_margins', 'asset_liability_ratio'])

  cols = list(metrics5years)
  cols.pop()
  cols.reverse()
  cols.append('ttm')
  metrics5years = metrics5years[cols]

  return {
    "cur_metrics" : pd.DataFrame([
      metrics.market_cap(ticker),
      metrics.cur_pe(ticker),
      metrics.cur_pfcf(ticker)
    ], index=['market_cap', 'cur_dilluted_pe', 'cur_pfcf'],
       columns=[datetime.now().strftime('%Y-%m-%d')]).T,

    "metrics5years" : metrics5years.T
  }

def write_to_workbook(folder, ticker, cur_metrics, five_year_metrics):
  wb = Workbook()
  try:
    ws = wb.active
    ws.title = 'To Date Fundamentals-Ratios'

    ws2 = wb.create_sheet("Five Year Fundamentals")

    for r in dataframe_to_rows(cur_metrics, index=True, header=True):
      ws.append(r)

    ws['B3'] = FORMAT_NUMBER_COMMA_SEPARATED1
    ws['C3'] = '0.000'
    ws['D3'] = '0.000'

    for r in dataframe_to_rows(five_year_metrics, index=True, header=True):
      ws2.append(r)

    for dim in ['B', 'C']:
      col = ws2.column_dimensions[dim]
      col.number_format = 'Currency'

    for dim in ['D', 'E']:
      col = ws2.column_dimensions[dim]
      col.number_format = FORMAT_CURRENCY_USD

    wb.save(f"{folder}/{ticker}.xlsx")
  finally:
    wb.close()

if __name__ == '__main__':
  parser = argparse.ArgumentParser(description="Displays EverythingMoney's 8 pillars for analyzing companies.")
  parser.add_argument('tickers', nargs='+', help='Tickers to analyze')
  args = parser.parse_args()

  folder = 'stock_reports'
  os.makedirs(folder, exist_ok=True)

  for ticker in args.tickers:
    report = analyze_ticker(ticker)
    write_to_workbook(folder, ticker, report['cur_metrics'], report['metrics5years'])
    sleep(60) # Alpha vantage allows only 5 api requests per minute thus timeout after each stock
